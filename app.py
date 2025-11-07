from flask import Flask, render_template, request, jsonify
import openpyxl
import os
import json

app = Flask(__name__)

@app.route('/')
def dashboard():
    return render_template('index.html')

@app.route('/cycletime', methods=['GET', 'POST'])
def cycletime():
    data_dir = os.getenv('CYCLETIME_DATA_DIR', 'data')
    files = [f for f in os.listdir(data_dir) if f.endswith('.xlsx')]
    headers = []
    data = []
    selected_file = None

    if request.method == 'POST':
        selected_file = request.form.get('file')
        if selected_file and selected_file in files:
            workbook = openpyxl.load_workbook(os.path.join(data_dir, selected_file))
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                if not headers:
                    headers = list(row)
                else:
                    data.append(list(row))

    return render_template('cycletime.html', files=files, headers=headers, data=data, selected_file=selected_file)

@app.route('/faultdelay', methods=['GET', 'POST'])
def faultdelay():
    data_dir = os.getenv('FAULTDELAY_DATA_DIR', 'faultdelay_data')
    files = [f for f in os.listdir(data_dir) if f.endswith('.xlsx')]
    headers = []
    data = []
    selected_file = None
    secondary_headers = []
    secondary_data = []

    if request.method == 'POST':
        selected_file = request.form.get('file')
        if selected_file and selected_file in files:
            workbook = openpyxl.load_workbook(os.path.join(data_dir, selected_file))
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                if not headers:
                    headers = list(row)
                else:
                    data.append(list(row))

            # Load secondary data
            secondary_dir = os.getenv('FAULTDELAY_DATA_DIR', 'stationfaultdelay_data')
            secondary_file_path = os.path.join(secondary_dir, selected_file)
            if os.path.exists(secondary_file_path):
                workbook_secondary = openpyxl.load_workbook(secondary_file_path)
                sheet_secondary = workbook_secondary.active
                for row in sheet_secondary.iter_rows(values_only=True):
                    if not secondary_headers:
                        secondary_headers = list(row)
                    else:
                        secondary_data.append(list(row))

    return render_template('faultdelay.html', files=files, headers=headers, data=data, selected_file=selected_file, secondary_headers=secondary_headers, secondary_data=secondary_data)

@app.route('/tipdress', methods=['GET', 'POST'])
def tipdress():
    data_dir = os.getenv('TIPDRESS_DATA_DIR', 'tipdress_data')
    files = [f for f in os.listdir(data_dir) if f.endswith('.xlsx')]
    headers = []
    data = []
    selected_file = None
    secondary_headers = []
    secondary_data = []

    if request.method == 'POST':
        selected_file = request.form.get('file')
        if selected_file and selected_file in files:
            workbook = openpyxl.load_workbook(os.path.join(data_dir, selected_file))
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                if not headers:
                    headers = list(row)
                else:
                    data.append(list(row))

            # Load secondary data
            secondary_dir =  os.getenv('TIPDRESS_DATA_DIR', 'tipdress_data')
            secondary_file_path = os.path.join(secondary_dir, selected_file)
            if os.path.exists(secondary_file_path):
                workbook_secondary = openpyxl.load_workbook(secondary_file_path)
                sheet_secondary = workbook_secondary.active
                for row in sheet_secondary.iter_rows(values_only=True):
                    if not secondary_headers:
                        secondary_headers = list(row)
                    else:
                        secondary_data.append(list(row))

    return render_template('tipdress.html', files=files, headers=headers, data=data, selected_file=selected_file, secondary_headers=secondary_headers, secondary_data=secondary_data)

@app.route('/dashboard-data')
def dashboard_data():
    json_path = os.getenv('DASHBOARD_JSON_PATH','parameters.json')
    if not json_path:
        return jsonify({'error': 'DASHBOARD_JSON_PATH environment variable not set'}), 500
    try:
        with open(json_path, 'r') as f:
            data = json.load(f)
        return jsonify(data)
    except FileNotFoundError:
        return jsonify({'error': 'JSON file not found'}), 500
    except json.JSONDecodeError:
        return jsonify({'error': 'Invalid JSON format'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)
